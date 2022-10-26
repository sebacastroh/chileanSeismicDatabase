import os
import openpyxl
import numpy as np
import pandas as pd
import multiprocessing
import scipy.io as spio
from myCodes import SeismicLibrary

# import time

path = os.getcwd()
Tn = np.array([0, 0.01, 0.02, 0.022, 0.025, 0.029, 0.03, 0.032, 0.035, 0.036,\
               0.04, 0.042, 0.044, 0.045, 0.046, 0.048, 0.05, 0.055, 0.06,\
               0.065, 0.067, 0.07, 0.075, 0.08, 0.085, 0.09, 0.095, 0.1,\
               0.11, 0.12, 0.13, 0.133, 0.14, 0.15, 0.16, 0.17, 0.18, 0.19,\
               0.2, 0.22, 0.24, 0.25, 0.26, 0.28, 0.29, 0.3, 0.32, 0.34, 0.35,\
               0.36, 0.38, 0.4, 0.42, 0.44, 0.45, 0.46, 0.48, 0.5, 0.55, 0.6,\
               0.65, 0.667, 0.7, 0.75, 0.8, 0.85, 0.9, 0.95, 1, 1.1, 1.2, 1.3,\
               1.4, 1.5, 1.6, 1.7, 1.8, 1.9, 2, 2.2, 2.4, 2.5, 2.6, 2.8, 3, 3.2,\
               3.4, 3.5, 3.6, 3.8, 4, 4.2, 4.4, 4.6, 4.8, 5, 5.5, 6, 6.5, 7, 7.5, 8, 8.5, 9, 9.5, 10])

xis = [0.02, 0.03, 0.05, 0.08, 0.1, 0.15, 0.2, 0.3, 0.5]

theta = np.linspace(0., np.pi, 181)

ct = np.cos(theta)[:,np.newaxis]
st = np.sin(theta)[:,np.newaxis]

def computeFlatFile(filename):
    # t1 = time.time()
    event = spio.loadmat(os.path.join(path, 'events_mat_corrected_v2', filename), struct_as_record=False, squeeze_me=True)
    event.pop('__header__')
    event.pop('__version__')
    event.pop('__globals__')

    Sx = {}
    Sy = {}
    Sz = {}
    GM = {}
    RotD0 = {}
    RotD50 = {}
    RotD100 = {}

    for xi in xis:
        Sx[xi] = []
        Sy[xi] = []
        Sz[xi] = []
        GM[xi] = []
        RotD0[xi] = []
        RotD50[xi] = []
        RotD100[xi] = []

    for key in sorted(event.keys()):
        station = event[key]
        arot = station.acc_1*ct + station.acc_2*st
        for xi in xis:
            spectrum = []
            for ax in arot:
                spectrum.append(SeismicLibrary.Spectrum(ax, station.dt, Tn, xi))
            Sx[xi].append(spectrum[0])
            Sy[xi].append(spectrum[90])
            Sz[xi].append(SeismicLibrary.Spectrum(station.acc_3, station.dt, Tn, xi))
            GM[xi].append(np.sqrt(spectrum[0]*spectrum[90]))
            RotD0[xi].append(np.min(spectrum, axis=0))
            RotD50[xi].append(np.median(spectrum, axis=0))
            RotD100[xi].append(np.max(spectrum, axis=0))

    for xi in xis:
        Sx[xi] = np.array(Sx[xi])
        Sy[xi] = np.array(Sy[xi])
        Sz[xi] = np.array(Sz[xi])
        GM[xi] = np.array(GM[xi])
        RotD0[xi] = np.array(RotD0[xi])
        RotD50[xi] = np.array(RotD50[xi])
        RotD100[xi] = np.array(RotD100[xi])

    return Sx, Sy, Sz, GM, RotD0, RotD50, RotD100

if __name__ == '__main__':

    filenames = sorted(os.listdir(os.path.join(path, 'events_mat_corrected_v2')))
    files_to_add = []
    for filename in filenames:
        if int(filename[:8]) >= 20210701:
            files_to_add.append(filename)
            
    pool = multiprocessing.Pool(24)
    results = pool.map(computeFlatFile, files_to_add)
    pool.close()
    
    wb = openpyxl.load_workbook('SpectralValues.xlsx', read_only=True)
    sheet = wb.worksheets[0]
    row_count = sheet.max_row
    
    columns = [str(T) for T in Tn.tolist()]
    columns[0] = 'PGA'
    sheet_names = ['Component 1', 'Component 2', 'Vertical component', 'Geometric mean', 'RotD0', 'RotD50', 'RotD100']
    filename = 'SpectralValues_2.xlsx'
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    writer.book = openpyxl.load_workbook('SpectralValues.xlsx')
    writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    for xi in xis:
        for i in range(7):
            sa = np.vstack([result[i][xi] for result in results])
            df = pd.DataFrame(sa/9.81, columns=columns)
            df.index = np.arange(row_count-1, row_count + len(df)-1)
            df.to_excel(writer, sheet_name=sheet_names[i] + ' - %0.2f' %xi, startrow=row_count, header=None)
    writer.save()

